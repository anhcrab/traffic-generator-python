import openpyxl

def get_all_user_agents():
    workbook = openpyxl.load_workbook("Storage/Traffic Generator.xlsx", True)
    user_agent_worksheet = workbook['User-Agents']
    user_agents = []
    total_rows = user_agent_worksheet.max_row
    for i in range(2, total_rows):
        user_agents.append({
            "id": user_agent_worksheet[f"A{i}"].value,
            "value": user_agent_worksheet[f"B{i}"].value,
            "device_type": user_agent_worksheet[f"C{i}"].value,
            "device_name": user_agent_worksheet[f"D{i}"].value,
        })
    workbook.close()
    return user_agents

def get_all_proxies():
    workbook = openpyxl.load_workbook("Storage/Traffic Generator.xlsx", True)
    proxy_worksheet = workbook['Proxies']
    proxies = []
    total_rows = proxy_worksheet.max_row
    for i in range(2, total_rows):
        proxies.append({
            "id": proxy_worksheet[f"A{i}"].value,
            "ip": proxy_worksheet[f"B{i}"].value,
            "port": proxy_worksheet[f"C{i}"].value,
            "code": proxy_worksheet[f"D{i}"].value,
            "country": proxy_worksheet[f"E{i}"].value,
            "anonymity": proxy_worksheet[f"F{i}"].value,
            "allow": get_boolean_value(proxy_worksheet[f"G{i}"].value),
            "google": get_boolean_value(proxy_worksheet[f"H{i}"].value),
            "https": get_boolean_value(proxy_worksheet[f"I{i}"].value),
        })
    workbook.close()
    return proxies

def get_all_traffic_urls():
    workbook = openpyxl.load_workbook("Storage/Traffic Generator.xlsx", True)
    traffics_worksheet = workbook['Traffic URLs']
    traffics = []
    total_rows = traffics_worksheet.max_row
    for i in range(2, total_rows):
        traffics.append({
            "id": traffics_worksheet[f"A{i}"].value,
            "keyword": traffics_worksheet[f"B{i}"].value,
            "url": traffics_worksheet[f"C{i}"].value,
            "rank": traffics_worksheet[f"D{i}"].value,
            "required_qty": traffics_worksheet[f"E{i}"].value,
            "current_qty": traffics_worksheet[f"F{i}"].value,
            "type": traffics_worksheet[f"G{i}"].value,
            "mobile": get_boolean_value(traffics_worksheet[f"H{i}"].value),
        })
    workbook.close()
    return traffics

def get_all_internal_links():
    workbook = openpyxl.load_workbook("Storage/Traffic Generator.xlsx", True)
    internal_links_worksheet = workbook['Internal links']
    internal_links = []
    total_rows = internal_links_worksheet.max_row
    for i in range(2, total_rows):
        if internal_links_worksheet[f"B{i}"].value is not None and internal_links_worksheet[f"C{i}"].value != '':
            internal_links.append(internal_links_worksheet[f"B{i}"].value)
    workbook.close()
    return internal_links

def handle_empty(value: str, replace: str):
    if value is None or value == "":
        return replace

def get_boolean_value(value: str):
    if value is None or value.lower() == "yes":
        return True
    else:
        return False

def get_boolean_str(value: str):
    if value is None:
        return ""
    if value:
        return "yes"
    return "no"

def set_proxy(proxy_id: int, key: str, value: str):
    workbook = openpyxl.load_workbook("Storage/Traffic Generator.xlsx")
    traffics_worksheet = workbook['Traffic URLs']
    keys = {
        "ip_address": "B",
        "port": "C",
        "code": "D",
        "country": "E",
        "anonymity": "F",
        "allow": "G",
        "google": "H",
        "https": "I",
    }
    traffics_worksheet[f"{keys[key]}{proxy_id + 1}"].value = value
    workbook.save("Storage/Traffic Generator.xlsx")
    workbook.close()

def set_traffic_url(traffic_id: int, key: str, value):
    workbook = openpyxl.load_workbook("Storage/Traffic Generator.xlsx")
    traffics_worksheet = workbook['Traffic URLs']
    keys = {
        "keyword": "B",
        "url": "C",
        "rank": "D",
        "required_qty": "E",
        "current_qty": "F",
    }
    traffics_worksheet[f"{keys[key]}{traffic_id + 1}"].value = value
    workbook.save("Storage/Traffic Generator.xlsx")
    workbook.close()

def get_traffic_url(traffic_id: int):
    for traffic in get_all_traffic_urls():
        if traffic["id"] == traffic_id:
            return traffic

class Storage:
    def get_all_traffic_urls(self):
        wb = openpyxl.load_workbook("Storage/Traffic Generator.xlsx", True)
